import openpyxl

# # 1. 파일 로드
# wb = openpyxl.load_workbook('long_data.xlsx')
# ws = wb.active
# new_wb = openpyxl.Workbook()
# new_ws = new_wb.active
# 엑셀을 읽어들이고 새로운 워크북과 워크시트를 생성합니다. 기존 데이터는 'long_data.xlsx' 파일에서 읽어오고, 변환된 데이터는 'wide_data_result.xlsx' 파일로 저장할 예정입니다.
#  

# 데이터를 키를 기준으로 정렬되었다고 가정하고, A열에 키, B열에 값이 있다고 가정합니다.
# 예시 데이터:
#  A열 (키) | B열 (값)
#  -------- | --------
#  key1     | value1
#  key1     | value2
#  key2     | value3
#  key2     | value4
#  key3     | value5
# 루프문을 통해서 키가 바뀔 때마다 새로운 행으로 이동하여 값을 추가하는 방식으로 Long to Wide 변환을 수행합니다.
# 결과 시트는 다음과 같이 될 것입니다:
#  A열 (키) | B열 (값1) | C열 (값2) | D열 (값3) | ...
#  -------- | --------- | --------- | --------- | ...
#  key1     | value1    | value2    |       | ...           
#  key2     | value3    | value4    |       | ...   
#  key3     | value5    |           |       | ...
# 1. 파일 로드 및 새로운 워크북 생성
# 2. 변수 초기화
#  - last_key: 이전 행의 키를 저장하는 변수
#  - current_row_idx: 결과 시트에서 현재 행 번호를 추적하는 변수
#  - current_col_idx: 결과 시트에서 현재 열 번호를 추적하는 변수    
# 3. 행 단위로 읽으며 Wide하게 펼치기
#  - 키가 같으면: 오른쪽(다음 열)으로 이동해서 값 추가
#  - 키가 다르면: 다음 행으로 내려가서 키와 첫 번째 값 작성
# 4. 결과 저장  
# 전체적으로, 이 코드는 Long 형식의 데이터를 Wide 형식으로 변환하는 과정을 보여줍니다. 각 키에 대해 여러 값이 있을 수 있으며, 이 값들은 같은 행의 오른쪽으로 추가됩니다. 키가 바뀔 때마다 새로운 행으로 이동하여 다음 키와 그에 해당하는 값을 작성합니다. 마지막으로, 변환된 데이터를 새로운 엑셀 파일로 저장합니다.

# 1. 파일 로드 및 새로운 워크북 생성
wb = openpyxl.load_workbook('long_data.xlsx')
ws = wb.active
new_wb = openpyxl.Workbook()
new_ws = new_wb.active  


# 2. 변수 초기화
last_key = None
current_row_idx = 0  # 결과 시트의 행 번호
current_col_idx = 2  # 결과 시트에서 값이 추가될 열 번호

# 3. 행 단위로 읽으며 Wide하게 펼치기
for row in ws.iter_rows(min_row=2, values_only=True):
    key, value = row[0], row[1] # A열은 키, B열은 값이라고 가정

    if key == last_key:
        # 키가 같으면: 오른쪽(다음 열)으로 이동해서 값 추가
        current_col_idx += 1
        new_ws.cell(row=current_row_idx, column=current_col_idx, value=value)
    else:
        # 키가 다르면: 다음 행으로 내려가서 키와 첫 번째 값 작성
        current_row_idx += 1
        current_col_idx = 2
        new_ws.cell(row=current_row_idx, column=1, value=key)
        new_ws.cell(row=current_row_idx, column=2, value=value)
        last_key = key

# 4. 결과 저장
new_wb.save('wide_data_result.xlsx')
print("Long to Wide 변환 완료!")

